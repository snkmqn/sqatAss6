import pytest
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import Font


URL = "https://the-internet.herokuapp.com/login"

EXCEL_PATH = "testdata_login.xlsx"
SHEET_NAME = "login"
RESULTS_PATH = "testdata_login_results.xlsx"

RUN_RESULTS = {}


def norm(s) -> str:
    return (str(s) if s is not None else "").strip().lower()


def get_header_map(ws):
    headers = [norm(c.value) for c in ws[1]]
    return {h: idx for idx, h in enumerate(headers)}


def read_test_data():
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    hm = get_header_map(ws)

    required = ["test_id", "username", "password", "expected_text", "expected_result"]
    missing = [h for h in required if h not in hm]
    if missing:
        raise ValueError(f"Missing columns in Excel: {missing}. Found: {list(hm.keys())}")

    data = []
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        test_id = str(row[hm["test_id"]]).strip()
        username = str(row[hm["username"]]).strip()
        password = str(row[hm["password"]]).strip()
        expected_text = str(row[hm["expected_text"]]).strip()
        expected_result = str(row[hm["expected_result"]]).strip().upper()
        data.append((row_index, test_id, username, password, expected_text, expected_result))
    return data


def write_results_file():
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    headers = [norm(c.value) for c in ws[1]]

    def ensure_col(header_name: str) -> int:
        h = norm(header_name)
        if h in headers:
            return headers.index(h) + 1
        cell = ws.cell(row=1, column=len(headers) + 1, value=header_name)
        cell.font = Font(bold=True)
        headers.append(h)
        return len(headers)

    actual_result_col = ensure_col("actual_result")
    actual_text_col = ensure_col("actual_text")

    for row_index, result in RUN_RESULTS.items():
        ws.cell(row=row_index, column=actual_result_col, value=result.get("actual_result", ""))
        ws.cell(row=row_index, column=actual_text_col, value=result.get("actual_text", ""))

    wb.save(RESULTS_PATH)


@pytest.fixture(scope="session")
def driver():
    options = Options()
    options.add_argument("--incognito")
    options.add_argument("--disable-notifications")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")

    drv = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()),
        options=options
    )
    drv.maximize_window()
    yield drv
    drv.quit()


@pytest.fixture(autouse=True)
def prepare_state(driver):
    driver.delete_all_cookies()
    driver.get(URL)
    yield


@pytest.fixture(scope="session", autouse=True)
def finalize_report():
    yield
    write_results_file()
    print(f"\n✅ Results saved to: {RESULTS_PATH}")


@pytest.mark.parametrize("row_index,test_id,username,password,expected_text,expected_result", read_test_data())
def test_login_ddt(driver, row_index, test_id, username, password, expected_text, expected_result):
    passed = False
    actual_text = ""

    try:
        wait = WebDriverWait(driver, 6)

        user_el = wait.until(EC.visibility_of_element_located((By.ID, "username")))
        pass_el = wait.until(EC.visibility_of_element_located((By.ID, "password")))

        user_el.clear()
        user_el.send_keys(username)

        pass_el.clear()
        pass_el.send_keys(password)

        driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

        flash = wait.until(EC.visibility_of_element_located((By.ID, "flash")))
        actual_text = flash.text.replace("×", "").strip()

        text_matches = expected_text in actual_text

        if expected_result == "PASS":
            assert text_matches, f"Expected text not found. expected='{expected_text}', actual='{actual_text}'"
        elif expected_result == "FAIL":
            assert not text_matches, f"Expected text should NOT appear, but it did. expected='{expected_text}', actual='{actual_text}'"
        else:
            raise ValueError("expected_result must be PASS or FAIL")

        passed = True

    finally:
        status = "PASSED" if passed else "FAILED"
        print(f"[{status}] {test_id} | user={username} | expected_result='{expected_result}' | expected_text='{expected_text}' | actual='{actual_text}'")

        RUN_RESULTS[row_index] = {
            "actual_result": status,
            "actual_text": actual_text
        }
