import os
import pytest
from openpyxl import load_workbook

from selenium.webdriver import Remote
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.edge.options import Options as EdgeOptions

URL = "https://the-internet.herokuapp.com/login"
EXCEL_PATH = "testdata_login.xlsx"
SHEET_NAME = "login"

BROWSERSTACK_USERNAME = os.environ.get("BROWSERSTACK_USERNAME")
BROWSERSTACK_ACCESS_KEY = os.environ.get("BROWSERSTACK_ACCESS_KEY")
if not BROWSERSTACK_USERNAME or not BROWSERSTACK_ACCESS_KEY:
    raise RuntimeError("Set env vars BROWSERSTACK_USERNAME and BROWSERSTACK_ACCESS_KEY")

REMOTE_URL = f"https://{BROWSERSTACK_USERNAME}:{BROWSERSTACK_ACCESS_KEY}@hub-cloud.browserstack.com/wd/hub"


def _norm(s) -> str:
    return (str(s) if s is not None else "").strip().lower()


def read_test_data():
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    headers = [_norm(c.value) for c in ws[1]]
    hm = {h: i for i, h in enumerate(headers)}

    required = ["test_id", "username", "password", "expected_text"]
    missing = [h for h in required if h not in hm]
    if missing:
        raise ValueError(f"Missing columns in Excel: {missing}. Found: {list(hm.keys())}")

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        test_id = str(row[hm["test_id"]]).strip()
        username = str(row[hm["username"]]).strip()
        password = str(row[hm["password"]]).strip()
        expected_text = str(row[hm["expected_text"]]).strip()
        data.append((test_id, username, password, expected_text))
    return data


def make_bs_driver(browser_name: str):
    bstack_options = {
        "os": "Windows",
        "osVersion": "10",
        "buildName": "DDT Login CrossBrowser (Chrome+Edge)",
        "sessionName": f"Login DDT - {browser_name}",
        "video": True,
        "idleTimeout": 60,
        "consoleLogs": "info",
    }

    if browser_name.lower() == "chrome":
        options = ChromeOptions()
        options.set_capability("browserName", "Chrome")
        options.set_capability("browserVersion", "latest")
        options.set_capability("bstack:options", bstack_options)
        return Remote(command_executor=REMOTE_URL, options=options)

    if browser_name.lower() == "edge":
        options = EdgeOptions()
        options.set_capability("browserName", "Edge")
        options.set_capability("browserVersion", "latest")
        options.set_capability("bstack:options", bstack_options)
        return Remote(command_executor=REMOTE_URL, options=options)

    raise ValueError(f"Unsupported browser: {browser_name}")


@pytest.fixture(scope="function")
def driver(request):
    browser = request.param
    drv = make_bs_driver(browser)
    try:
        yield drv
    finally:
        drv.quit()


@pytest.mark.parametrize("driver", ["Chrome", "Edge"], indirect=True)
@pytest.mark.parametrize("test_id,username,password,expected_text", read_test_data())
def test_login_cross_browser(driver, test_id, username, password, expected_text):
    wait = WebDriverWait(driver, 3)

    driver.get(URL)

    wait.until(EC.visibility_of_element_located((By.ID, "username"))).send_keys(username)
    wait.until(EC.visibility_of_element_located((By.ID, "password"))).send_keys(password)

    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

    flash = wait.until(EC.visibility_of_element_located((By.ID, "flash")))
    actual_text = flash.text.replace("×", "").strip()

    assert expected_text in actual_text, f"{test_id} expected='{expected_text}' actual='{actual_text}'"
